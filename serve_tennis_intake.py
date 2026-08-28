"""Serve Tennis export -> scheduler inputs (events, seeds).

Roster-intake stage. Reads the WWTC player-list (.csv baseline; .xlsx also supported) and builds EventSpec
objects (singles + doubles) plus a seeds map, ready to drop into MultiConfig.
No engine changes: this is a pure front end that emits the same objects you'd
build by hand.

Format rule (confirmed): a division with 3..THRESHOLD entrants (singles) or
teams (doubles) -> round_robin; otherwise single_elim. 2-entry divisions are a
single final. Override any division explicitly via `format_overrides`.
"""
import csv
import re
from collections import OrderedDict
from openpyxl import load_workbook
from scheduler_multi import EventSpec, Team

RR_THRESHOLD = 5     # <= this many entrants/teams -> round robin (when >=3)


def _read_rows(path, sheet="Player List"):
    """Read a player-list export into (title, rows) where rows is a list of cell-value
    tuples, header first. Supports **.xlsx** (openpyxl) and **.csv** (comma-delimited,
    all-string cells; a UTF-8 BOM is stripped). CSV is friendlier to zips than openpyxl —
    leading zeros (e.g. New England 06105) survive as text. Raises loudly on an empty file."""
    if str(path).lower().endswith(".csv"):
        with open(path, newline="", encoding="utf-8-sig") as fh:
            rows = [tuple(r) for r in csv.reader(fh)]
        title = "csv"
    else:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet] if sheet in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(values_only=True))
        title = ws.title
    if not rows:
        raise ValueError(f"No rows found in {path!r} (sheet {sheet!r}). Is this the right file/sheet?")
    return title, rows


# ---- helpers ----------------------------------------------------------------
def _age(name):
    m = re.search(r"(\d+)\s*&\s*over", name or "")
    return int(m.group(1)) if m else 0

def _recovery_for(name):
    # FAC within-division minimums: 60-and-over -> 90 min, otherwise 60 min.
    return 90 if _age(name) >= 60 else 60

def _cap_for(name):
    # FAC Table 9 max matches/day, keyed to the division's age bracket. Age is
    # parsed from the division name via _age() — the same source _recovery_for uses.
    #   Adult age <=55 -> 6  (of which <=4 may be singles; that singles sub-cap is
    #                         NOT yet enforced — needs per-player match-type tracking,
    #                         logged as a known limitation in STATUS §9)
    #   age 60-80      -> 4
    #   age 85+        -> 3
    # (USTA adult brackets step 55->60 and 80->85, so the 56-59 / 81-84 gaps never
    #  occur; a division with no parseable age reads 0 and lands in the <=55 tier.)
    a = _age(name)
    if a >= 85:
        return 3
    if a >= 60:
        return 4
    return 6

def _is_doubles(name):
    return "doubles" in (name or "").lower()

def _int(v):
    try:
        return int(str(v).strip())
    except (TypeError, ValueError):
        return None

def _rank(v):
    # Coerce to a number for the unseeded sort. xlsx yields ints (unchanged ordering);
    # csv yields strings ("261" -> 261.0). Blank/None -> sorts last.
    try:
        return float(v)
    except (TypeError, ValueError):
        return 10**9


# ---- main -------------------------------------------------------------------
def load_export(path, sheet="Player List", threshold=RR_THRESHOLD,
                format_overrides=None, default_cap=None, status="Accepted"):
    """Return (events, seeds_by_event, warnings).

    events           : list[EventSpec], ready for MultiConfig(events=...)
    seeds_by_event   : {event_name: {team_label: seed_int}} for render_all
    warnings         : list[str] of anything skipped or coerced
    """
    fo = format_overrides or {}
    _title, rows = _read_rows(path, sheet)   # .xlsx or .csv; raises loudly on an empty file

    header = rows[0]
    hdr_w = len(header)
    ix = {h: i for i, h in enumerate(header)}

    # Ragged-row guard: in read_only mode openpyxl trims trailing empty cells, so
    # data rows can arrive narrower than the header. Check the row is long enough,
    # not just that the column name exists; missing trailing cells read as None.
    g = lambda r, c: r[ix[c]] if (c in ix and ix[c] < len(r)) else None
    full = lambda r: f"{g(r,'First Name')} {g(r,'Last Name')}".strip()

    # Loud failure #1 — required grouping column missing/unmapped. Without this,
    # every row groups under Event=None and the build silently yields one bogus
    # nameless division.
    if "Event" not in ix:
        raise ValueError(
            f"Required 'Event' column not found. Header has: {sorted(str(h) for h in ix)}. "
            f"Rename the division column to 'Event' (column aliasing is a later update)."
        )

    data = [r for r in rows[1:] if status is None or g(r, "Entry Status") == status]

    # Loud failure #2 — nothing survives the status filter. Without this, the build
    # silently returns zero divisions and the run looks 'successful' but empty.
    if not data:
        seen = sorted({str(g(r, "Entry Status")) for r in rows[1:]})
        raise ValueError(
            f"No entrants matched status={status!r}. 'Entry Status' values present: {seen}. "
            f"Pass a matching status=... or set status=None to disable the filter."
        )

    bydiv = OrderedDict()
    for r in data:
        bydiv.setdefault(g(r, "Event"), []).append(r)

    events, seeds_by_event, warnings = [], {}, []

    # Data-shape notice — surface ragged input so the TD knows the export wasn't
    # clean (trimmed trailing cells were read as blank, which is usually correct
    # but worth seeing). Advisory only; does not affect scheduling.
    ragged = sum(1 for r in data if len(r) < hdr_w)
    if ragged:
        warnings.append(
            f"{ragged} accepted row(s) narrower than the {hdr_w}-column header; "
            f"trailing cells read as blank (ragged export)."
        )
    for name, rs in bydiv.items():
        doubles = _is_doubles(name)
        ranked = []          # (Team, seed_or_None, rank_proxy)
        seedmap = {}

        if doubles:
            byname = {full(r): r for r in rs}
            used = set()
            for r in rs:
                me = full(r)
                if me in used:
                    continue
                p = g(r, "Partner")
                if not p or p not in byname:
                    warnings.append(f"{name}: '{me}' has no resolvable partner -> dropped")
                    continue
                pr = byname[p]
                used.update((me, p))
                t = Team(tid="+".join(sorted([str(g(r, "USTA ID")), str(g(pr, "USTA ID"))])),
                         members=[me, p])
                seed = _int(g(r, "Seed")) or _int(g(pr, "Seed"))
                if seed:
                    seedmap[t.label()] = seed
                ranked.append((t, seed, min(_rank(g(r, "Ranking")), _rank(g(pr, "Ranking")))))
        else:
            for r in rs:
                t = Team(tid=str(g(r, "USTA ID")), members=[full(r)])
                seed = _int(g(r, "Seed"))
                if seed:
                    seedmap[t.label()] = seed
                ranked.append((t, seed, _rank(g(r, "Ranking"))))

        if len(ranked) < 2:
            warnings.append(f"{name}: {len(ranked)} entr{'y' if len(ranked)==1 else 'ies'} -> skipped (no draw)")
            continue

        # seeded first (seed asc), then unseeded by ranking asc
        ranked.sort(key=lambda x: (0, x[1]) if x[1] else (1, x[2]))
        teams = [t for (t, _, _) in ranked]
        n = len(teams)

        if name in fo:
            fmt = fo[name]
            if fmt == "round_robin" and n < 3:
                warnings.append(f"{name}: override RR but only {n} teams -> using single_elim")
                fmt = "single_elim"
        else:
            fmt = "round_robin" if 3 <= n <= threshold else "single_elim"

        events.append(EventSpec(
            name=name, fmt=fmt, teams=teams,
            match_minutes=90,
            recovery_minutes=_recovery_for(name),
            precedence=1 if doubles else 0,     # singles before doubles
            # FAC Table 9 age-based cap (default); an explicit default_cap forces a
            # flat cap across all divisions for back-compat / testing.
            max_matches_per_day=_cap_for(name) if default_cap is None else default_cap,
        ))
        if seedmap:
            seeds_by_event[name] = seedmap

    return events, seeds_by_event, warnings


if __name__ == "__main__":
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else "/mnt/user-data/uploads/WWTC_Sample_Player_List.csv"
    events, seeds, warns = load_export(path)
    rr = sum(1 for e in events if e.fmt == "round_robin")
    print(f"built {len(events)} divisions  ({rr} round_robin, {len(events)-rr} single_elim)")
    for e in events:
        sd = len(seeds.get(e.name, {}))
        print(f"  {e.fmt:<12} {len(e.teams):>3} teams  rec={e.recovery_minutes}  seeds={sd}  {e.name}")
    if warns:
        print("\nwarnings:")
        for w in warns:
            print("  -", w)
