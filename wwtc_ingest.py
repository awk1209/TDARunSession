"""wwtc_ingest.py — R1b: join the real WWTC player lists and resolve the finalized draw.

Real-data ingest for the MVP rescope (tournament_desk plane). Two Apple Numbers player-list files
per level, joined on **USTA ID**, then the finalized bracket (from `draws_pdf.parse_draws`) is
resolved to full player identities:

  - **TD list** (Tournament Desk): Name, Gender, **Events** (the division-name authority — its
    strings match the draws PDF), City/State, **Draw status**, USTA ID, WTN. Identified by a
    `Draw status` column.
  - **Serve Tennis list "w/ sections"**: adds **City / District / Section**, Year of birth,
    Selection. Identified by a `Section` column. Joins onto TD on USTA ID (ST ⊆ TD).

Locality has **no zip** in the real export → it is **Section / City** from the ST list (the Q1
answer). The canonical player `name` is built as `f"{First} {Last}".strip()` — identical to
`serve_tennis_intake.load_export` / `roster.normalize`, so it is the same human-identity key the
engine uses (`Team.members` / `Match.humans`).

This supersedes `roster.py`'s `TournamentDeskAdapter` stub; the serve_tennis CSV path stays the
legacy sample path. Read-only; no engine coupling (R1c builds EventSpecs from `resolve_draws`).
"""
from __future__ import annotations

import glob
import os
import re
from dataclasses import dataclass, field
from typing import Optional

import draws_pdf
import field_source


def read_numbers(path):
    """(header, rows) from the first table of a .numbers workbook (via numbers-parser)."""
    from numbers_parser import Document
    doc = Document(path)
    table = doc.sheets[0].tables[0]
    rows = table.rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    return header, rows[1:]


def read_xlsx(path):
    """(header, rows) from the first worksheet of an .xlsx workbook (via openpyxl). Same shape as
    read_numbers: header = list[str]; rows = list of positional value-lists. Fully-empty rows are
    dropped so trailing padding doesn't masquerade as entrants."""
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    rows = list(wb.worksheets[0].iter_rows(values_only=True))
    wb.close()
    if not rows:
        return [], []
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    body = [list(r) for r in rows[1:] if any(c is not None for c in r)]
    return header, body


def read_csv(path):
    """(header, rows) from a CSV player list. Same shape as read_numbers/read_xlsx: header =
    list[str]; rows = list of positional value-lists, fully-empty rows dropped.

    DESK-1 (run report D3, 2026-08-09). CSV is a first-class Tournament Desk export format and
    the ingest did not read it: `read_table` dispatched `.xlsx` | `.numbers` only, so a valid CSV
    player list reported **"not found" with `candidates: []`** and hard-blocked at Step 4 — after
    the finals map is already done. The 2027 mock run got past it by hand-converting the file.
    A BOM is stripped (`utf-8-sig`) because that is what Excel writes, and an unstripped BOM
    would corrupt the FIRST header cell only — which is `First Name` on the TD list, so the
    join would silently lose every name rather than fail loudly."""
    import csv as _csv
    with open(path, newline="", encoding="utf-8-sig") as fh:
        rows = [r for r in _csv.reader(fh)]
    if not rows:
        return [], []
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    body = [list(r) for r in rows[1:] if any(str(c).strip() for c in r)]
    return header, body


def read_table(path):
    """(header, rows) from a player-list workbook — dispatch by extension
    (.xlsx | .numbers | .csv)."""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xlsx":
        return read_xlsx(path)
    if ext == ".numbers":
        return read_numbers(path)
    if ext == ".csv":
        return read_csv(path)
    raise ValueError(f"unsupported player-list format {ext!r}: {path}")


def _classify(path):
    """('td'|'st'|None, level) from a player-list file's header + filename."""
    try:
        header, _ = read_table(path)
    except Exception:
        return None, None
    hset = set(header)
    kind = "st" if "Section" in hset else ("td" if "Draw status" in hset else None)
    m = re.search(r"\bL(\d)\b", os.path.basename(path)) or re.search(r"_L(\d)_", os.path.basename(path))
    level = m.group(1) if m else None
    return kind, level


def resolve_player_lists(level="2"):
    """Find the TD + ST .numbers pair for a level across the ephemeral upload dirs.
    Env overrides: $WWTC_TD_LIST / $WWTC_ST_LIST."""
    td = os.environ.get("WWTC_TD_LIST")
    st = os.environ.get("WWTC_ST_LIST")
    if td and st and os.path.exists(td) and os.path.exists(st):
        return td, st
    found = {"td": None, "st": None}
    data_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "wwtc-2026")
    for root in (data_dir, "/mnt/user-data/uploads", "/root/.claude/uploads"):
        for pat in ("*.xlsx", "*.numbers", "*.csv"):     # DESK-1 (D3): CSV is a desk export too
            for p in sorted(glob.glob(os.path.join(root, "**", pat), recursive=True)):
                kind, lvl = _classify(p)
                if kind and (lvl == level or lvl is None) and not found[kind]:
                    found[kind] = p
    if not (found["td"] and found["st"]):
        raise FileNotFoundError(
            f"Could not resolve TD+ST player lists for L{level}. "
            f"Found: {found}. Set $WWTC_TD_LIST / $WWTC_ST_LIST.")
    return found["td"], found["st"]


_TRUNC = re.compile(r"\s*(\.\.\.|…)\s*$")


def _norm(s):
    """Casefold + collapse whitespace + drop periods, for name matching."""
    return re.sub(r"\s+", " ", str(s or "").replace(".", "").strip()).casefold()


def _core(display):
    """(normalized-core, is_truncated). A trailing '...'/'…' marks a name the PDF cut off."""
    raw = str(display or "").strip()
    trunc = bool(_TRUNC.search(raw))
    return _norm(_TRUNC.sub("", raw)), trunc


@dataclass
class Player:
    usta_id: str
    name: str                 # "First Last" — engine human key
    first: str = ""
    last: str = ""
    gender: str = ""
    city: str = ""
    district: str = ""
    section: str = ""
    state: str = ""
    yob: Optional[int] = None
    wtn_singles: Optional[float] = None
    wtn_doubles: Optional[float] = None
    events: list = field(default_factory=list)     # divisions entered (from TD Events)
    td_status: str = ""
    raw_td: dict = field(default_factory=dict)
    raw_st: dict = field(default_factory=dict)
    # ROSTER-1: the desk's Draw status split per division — {division: status | None}, None =
    # the row's two cells disagree in length so nothing positional can be trusted. Keys are the
    # row's DISTINCT divisions, so this doubles as the canonical person-division entry set.
    entry_status: dict = field(default_factory=dict)
    # ROSTER-1: the originating TD row per division. A player entered at BOTH levels has two TD
    # rows with two different `Draw status` cells; the union keeps one Player, so the per-division
    # row is the only way an output can quote the right level's words back to the TD.
    raw_td_by_division: dict = field(default_factory=dict)


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _dedup(seq):
    """Order-preserving de-duplication — first occurrence wins."""
    out, seen = [], set()
    for x in seq:
        if x not in seen:
            seen.add(x)
            out.append(x)
    return out


# DESK-1 (2026-08-09): the third value `entry_status` can carry. `None` means UNRESOLVED — the
# row is ambiguous and no rule may allocate it (ruling (b): print the raw row, invent nothing).
# `NO_DESK_WORD` means RESOLVED WITH NOTHING TO SAY: the desk gave this row exactly one status
# word, that word is not a non-playing one, and the Raw Draw does not carry this entry — so the
# fact is settled (they are not playing it) and the desk simply never wrote a reason. It is a
# quiet row, NOT a flagged one, which is the whole difference between the 12 rows it covers on
# the 2026 field and the 6 that stay open.
NO_DESK_WORD = ""

# `_entry_row`'s "the caller did not pass a status" marker. It cannot be None: None is a real,
# load-bearing value there (the row is ambiguous and flagged).
_UNSET = object()


def _non_playing(tokens):
    """The row's non-playing status words, order-preserving and deduped. `Selected` is the only
    word that means "playing"; everything else the desk writes (`Withdrawn`, `Alternate`,
    `Unpaired`, `Selected - ALT`) is a reason someone is not in a draw."""
    out = []
    for t in tokens:
        if t.casefold() != "selected" and t not in out:
            out.append(t)
    return out


def entry_status_map(events_raw, status_raw, drawn=None):
    """{division: status | NO_DESK_WORD | None} for one TD row. None = unresolved (ambiguous).

    `drawn` (DESK-1, 2026-08-09) is the set of `_norm_div` keys this person is in the **Raw
    Draw** of. Pass it and the map is read the ruled way (below); pass None — every caller that
    has no draws behind it, `load_players` included — and the shipped positional read is
    returned unchanged, so a synthetic caller sees exactly what it always saw.

    **THE DRAW-INFORMED READING** (Operator ruling 7.6, 2026-08-08). The desk cannot export a
    per-entry status (question closed 8/8), so the entry list's one `Draw status` cell has to be
    allocated across the row's entries by rule. The draw is what allocates it:

      1. **Draw membership wins.** A division on this row that is in the person's Raw Draw reads
         `Selected`, whatever the entry list says — the draw is truth (Director, 2026-07-27).
         This is what makes a non-playing word on a drawn division unreachable: **12 entries on
         the 2026 field** read `Unpaired` / `Withdrawn` / `Selected - ALT` for a division the
         person is standing in, and all 12 now read `Selected`.
      2. **Divisions NOT in the Raw Draw take the row's non-playing words**, and how depends on
         whether the two cells can be trusted against each other:
         a. **counts match** — the positional pairing is real data the desk actually wrote, so
            it is kept, FIRST-WINS on a repeated token (hazard 1 below). Throwing it away would
            turn 7 people whose withdrawal the desk stated plainly into ambiguous pool rows.
         b. **counts disagree** — nothing positional can be trusted (hazard 2 below), so the row
            is read as a whole: ONE distinct non-playing word covers every not-drawn entry; NO
            non-playing word gives `NO_DESK_WORD`; TWO OR MORE stays `None` and flagged, because
            no rule can say which division withdrew and which is an alternate.

    On the 2026 field the 44 entries this function leaves unresolved today read **38 resolved ·
    6 open · 0 contradicted by the draw** — the 6 being two people whose single row carries
    three status words (`Selected, Withdrawn, Alternate` and `Unpaired, Withdrawn, Alternate`).

    Two hazards, both live on the 2026 field, and both now have an EXPLICIT case above rather
    than an inherited one:

      - **The cell repeats division tokens** (12 rows, 12 people, 25 entries). A repeated
        doubles token is one person entered in one division with TWO partnerships, and the desk
        writes a word for each. The pair must be formed FIRST and deduped after, so the first
        status for a division wins. Deduping before the pair is formed — or letting a
        `{e: sts[i] for i, e in enumerate(evs)}` comprehension collide on the repeated key —
        silently promotes the LAST status, which turned **6 `Withdrawn` entries into
        available-looking `Alternate`s**. `_division_index` documents the same hazard for the
        candidate pool; this is its status-side twin.
        **The draw resolves 5 of the 12 outright** — Shea, Irwin, Voros, Dobson and (on its
        other token) Harmon are IN the doubles draw they entered twice, so rule 1 reads them
        `Selected` and first-wins never runs. The other 7 keep first-wins under 2a, which is why
        `Withdrawn` still beats `Alternate` for them and none of them drift into the pool.

      - **The two counts disagree** (266 rows across both levels; 44 of the resulting unresolved
        entries, across 35 people, land in no draw and so reach the exceptions list — the rest are
        drawn, where status is never consulted). Nothing positional can be trusted, so every
        division on the row is left UNRESOLVED.
        **Operator ruling (b), 2026-07-31:** print the raw row rather than invent a
        per-division status. These people are handled externally in Tournament Desk; the ambiguity
        belongs on the report, not in a courier stop.

    Status is NEVER a placement input (Director, 2026-07-27: the draw is truth) — this only labels
    people who are in no draw. **The keys never move**: they are `_dedup(events_raw)` on every
    path, which is what keeps `reconcile_entries`'s accounting closed when the reading changes.
    """
    evs = list(events_raw or [])
    keys = _dedup(evs)
    sts = [s.strip() for s in str(status_raw or "").split(",") if s.strip()]
    counts_match = len(sts) == len(evs)

    positional = {}
    if counts_match:
        for e, s in zip(evs, sts):
            positional.setdefault(e, s)           # first status for a repeated token wins

    if drawn is None:                             # the shipped read — no draws behind the caller
        return dict(positional) if counts_match else {e: None for e in keys}

    nonplay = _non_playing(sts)
    out = {}
    for e in keys:
        if _norm_div(e) in drawn:
            out[e] = "Selected"                   # 1 · the draw is truth
        elif counts_match:
            # 2a · the desk's own pairing, first-wins. A `Selected` here is the desk saying the
            # player is IN a division the Raw Draw does not carry them in — a playing word on a
            # not-playing entry, which is no reason at all. It reads as settled-with-no-reason,
            # exactly like the 2b case below, so that **`Selected` can never appear on a row the
            # tool is explaining an absence for**. 1 entry on the 2026 field (Patrick Crow,
            # Men's 60 & over doubles) — the single row the retired review trigger used to fire on.
            w = positional.get(e)
            out[e] = NO_DESK_WORD if (w and w.casefold() == "selected") else w
        elif len(nonplay) == 1:
            out[e] = nonplay[0]                   # 2b · one word covers the row
        elif not nonplay:
            out[e] = NO_DESK_WORD                 # 2b · settled, but the desk gave no reason
        else:
            out[e] = None                         # 2b · ruling (b) — open, and flagged
    return out


def _td_rows(p):
    """The player's TD rows, one per level, as (raw, divisions, events_raw, status_cell).

    A person entered at BOTH levels has two rows with two different `Draw status` cells, and the
    reading is per ROW — `raw_td_by_division` is the only thing that says which entry belongs to
    which cell, so it is grouped by row identity here rather than by level. Rows carrying no
    `Events` cell are skipped: synthetic callers build `Player.entry_status` directly and there
    is nothing to re-read for them."""
    groups = {}
    for div, raw in p.raw_td_by_division.items():
        groups.setdefault(id(raw), (raw, []))[1].append(div)
    for raw, divs in groups.values():
        evs = [e.strip() for e in str((raw or {}).get("Events") or "").split(",") if e.strip()]
        if evs:
            yield raw, divs, evs, str((raw or {}).get("Draw status") or "").strip()


def resolved_entry_status(p, drawn):
    """`p.entry_status` re-read against the Raw Draw (DESK-1). `drawn` is the set of `_norm_div`
    keys this person is in a draw of.

    **The entry SET never changes** — the result is keyed on `p.entry_status` and any division a
    row cannot account for keeps the value it already had. That is deliberate: the closed
    accounting (`entries == scheduled + exceptions`) is keyed on this dict's length, so a reading
    that could add or drop a key would break the one invariant this reallocation must not touch.
    The Player is never mutated; the caller owns the result."""
    read = {}
    for _raw, _divs, evs, cell in _td_rows(p):
        read.update(entry_status_map(evs, cell, drawn=drawn))
    return {d: read.get(d, p.entry_status[d]) for d in p.entry_status}


def unpaired_partners(p, drawn, partners):
    """{division: partner name} for every entry the desk marked `Unpaired` that is standing in
    that division's Raw Draw WITH a partner (run report D6, 2026-08-09).

    The failure this finds is a person the entry list says has nobody to play with, on a doubles
    court with somebody. Nothing in the tool reported it: the two sources were never compared,
    and after DESK-1's own reading the entry reads `Selected` (rule 1 — the draw is truth), so
    the `Unpaired` word is gone from the status by the time any surface could notice it. The
    check therefore reads the **desk's raw cell**, not the resolved status, and names both
    sources so the desk can say which one is stale.

    Which entry the word lands on follows the reading: positionally when the counts match, and
    otherwise every entry on the row, because a row whose counts disagree cannot say. **16
    entries for 14 people on the 2026 field**; the 2027 mock run reported 37 of 39."""
    out = {}
    for _raw, _divs, evs, cell in _td_rows(p):
        if "unpaired" not in cell.casefold():
            continue
        sts = [s.strip() for s in cell.split(",") if s.strip()]
        if len(sts) == len(evs):
            pos = {}
            for e, s in zip(evs, sts):
                pos.setdefault(e, s)
            targets = [e for e, s in pos.items() if "unpaired" in s.casefold()]
        else:
            targets = _dedup(evs)
        for d in targets:
            mate = (partners or {}).get((p.name, _norm_div(d)))
            if _norm_div(d) in drawn and mate:
                out[d] = mate
    return out


def load_players(td_path=None, st_path=None, level="2"):
    """Join the TD + ST lists on USTA ID → {usta_id: Player}. TD is the authority for identity
    + division membership; ST enriches with Section/District/City/YoB.

    S-2 SEAM (Operator ruling, 2026-08-24 — option 1). The other half of the ingest boundary.
    A projected field carries its own matched roster — last season's real people plus one
    placeholder per slot the roster genuinely could not fill — and serving it here is what
    makes the pair MATCHED rather than a set of draws full of names nothing can resolve. It is
    also what retires the two stubs the scaffolding needed: with a roster to reconcile against,
    the name check and the entered-but-not-drawn list run for real. With nothing installed the
    body below runs exactly as it did before the seam existed, byte for byte.
    """
    projected = field_source.installed()
    if projected is not None:
        return projected.players_for(level)
    if td_path is None or st_path is None:
        td_path, st_path = resolve_player_lists(level)
    th, trows = read_table(td_path)
    sh, srows = read_table(st_path)
    ti = {h: i for i, h in enumerate(th)}
    si = {h: i for i, h in enumerate(sh)}

    def g(row, ix, col):
        return row[ix[col]] if (col in ix and ix[col] < len(row)) else None

    def uid(v):
        if v is None:
            return None
        try:
            return str(int(v))
        except (TypeError, ValueError):
            return str(v).strip()

    # Serve Tennis enrichment, keyed by USTA ID (column "ID")
    st_by_id = {}
    for r in srows:
        k = uid(g(r, si, "ID"))
        if k:
            st_by_id[k] = r

    players = {}
    for r in trows:
        k = uid(g(r, ti, "USTA ID"))
        if not k:
            continue
        first = str(g(r, ti, "First Name") or "").strip()
        last = str(g(r, ti, "Last Name") or "").strip()
        events = [e.strip() for e in str(g(r, ti, "Events") or "").split(",") if e.strip()]
        p = Player(
            usta_id=k, name=f"{first} {last}".strip(), first=first, last=last,
            gender=str(g(r, ti, "Gender") or "").strip(),
            city=str(g(r, ti, "City") or "").strip(), state=str(g(r, ti, "State") or "").strip(),
            wtn_singles=_num(g(r, ti, "WTN Singles")), wtn_doubles=_num(g(r, ti, "WTN Doubles")),
            events=events, td_status=str(g(r, ti, "Draw status") or "").strip(),
            raw_td={th[i]: (r[i] if i < len(r) else None) for i in range(len(th))},
        )
        p.entry_status = entry_status_map(events, p.td_status)          # ROSTER-1
        p.raw_td_by_division = {e: p.raw_td for e in p.entry_status}    # ROSTER-1
        sr = st_by_id.get(k)
        if sr is not None:
            p.district = str(g(sr, si, "District") or "").strip()
            p.section = str(g(sr, si, "Section") or "").strip()
            if not p.city:
                p.city = str(g(sr, si, "City") or "").strip()
            # DESK-1 (D3, 2026-08-09): a workbook reader hands back a NUMBER and the CSV reader
            # hands back the string `"1975"`, so an `isinstance(int, float)` test alone drops
            # every birth year on the CSV lane silently — the lane would read as equivalent and
            # quietly carry one field less. Parsed by value, not by the reader's type.
            yob = g(sr, si, "Year of birth")
            yob = _num(yob)
            p.yob = int(yob) if yob is not None else None
            p.raw_st = {sh[i]: (sr[i] if i < len(sr) else None) for i in range(len(sh))}
        players[k] = p
    return players


LEVELS = ("1", "2")


def load_players_combined(levels=LEVELS):
    """Union of the levels' player records, keyed by USTA ID — **one record per human**.

    ROSTER-1 (2026-07-31). L1 and L2 are one tournament on shared courts and **84 people are
    entered at both levels**, each with their own TD row. A plain `dict.update()` — what
    `wwtc_pipeline.build_combined` did — replaces the L1 record with the L2 one wholesale, and
    the L1 `events` go with it: **85 entries erased, 9 of them for people in no draw at all**
    (8 people), who then vanish from every reconciliation instead of appearing on the exceptions
    list. It also blinded `resolve_draws` in exactly the four L1 Mixed divisions, which is why
    `avoidance.py` and `schedule_report.py` each grew a local repair; both now call this instead,
    so there is **one** union rather than three.

    What merges and what does not:
      - `events` / `entry_status` / `raw_td_by_division` are **unioned** across levels — these are
        per-division facts and each level holds a real, non-overlapping part of them.
      - every scalar (name, city, section, WTN, `td_status`, `raw_td`) keeps the **last** level's
        value, exactly as `dict.update()` produced. `roster_meta` therefore returns byte-identical
        output, which is what keeps placement unmoved: locality and the engine's human keys do not
        move because this function changed.
    """
    players = {}
    for lvl in levels:
        for uid, p in load_players(level=lvl).items():
            prior = players.get(uid)
            if prior is not None:
                p.events = _dedup(list(prior.events) + list(p.events))
                p.entry_status = {**prior.entry_status, **p.entry_status}
                p.raw_td_by_division = {**prior.raw_td_by_division, **p.raw_td_by_division}
            players[uid] = p
    return players


# ---- draw resolution -------------------------------------------------------

@dataclass
class Resolved:
    """One drawn entrant resolved to canonical player(s)."""
    division: str
    kind: str                 # "slot" | "rr"
    pos: Optional[int]        # elim slot position (None for RR)
    display: str
    seed: Optional[int]
    is_bye: bool = False
    player_ids: list = field(default_factory=list)   # 1 (singles) or 2 (doubles) USTA IDs
    ok: bool = False
    note: str = ""
    ref: str = ""             # stable id within its division ("slot-3" / "Group 1#0")
    group: str = ""           # RR group name ("" for elim)
    is_doubles: bool = False


def _division_index(players):
    """division -> list[Player] entered in it (from TD Events). Deduped by USTA ID: the TD Events
    cell repeats some division tokens, which would otherwise double-count a player in the pool and
    trip spurious same-surname ambiguity flags."""
    idx, seen = {}, {}
    for p in players.values():
        for ev in set(p.events):                       # set() drops duplicate division tokens
            key = (ev, p.usta_id)
            if key in seen:
                continue
            seen[key] = True
            idx.setdefault(ev, []).append(p)
    return idx


def _match_singles(display, cands, used):
    d, trunc = _core(display)
    avail = [p for p in cands if p.usta_id not in used]
    exact = [p for p in avail if _norm(p.name) == d]
    if len(exact) == 1:
        return exact[0], ""
    if len(exact) > 1:
        exact.sort(key=lambda p: p.usta_id)
        return exact[0], f"ambiguous singles '{display}' ({len(exact)}) -> {exact[0].name}"
    # prefix: RR display carries a city tail, or the PDF truncated the name ("...")
    pref = [p for p in avail if d.startswith(_norm(p.name)) or (trunc and _norm(p.name).startswith(d))]
    if pref:
        pref.sort(key=lambda p: -len(p.name))     # most specific (longest) name wins
        note = f"{'truncated' if trunc else 'prefix'}-matched '{display}' -> {pref[0].name}"
        return pref[0], ("" if _norm(pref[0].name) == d else note)
    return None, f"no singles match for '{display}'"


def _city_pick(cands, display):
    """Among same-surname candidates, keep those whose city appears in the entrant display
    (the RR grid carries partner cities). Returns the narrowed list (or the input if no signal)."""
    d = _norm(display)
    hits = [p for p in cands if p.city and _norm(p.city) in d]
    return hits if len(hits) == 1 else cands


def _match_last(surname, cands, used, display="", allow_ambiguous=True):
    """Resolve one doubles partner surname to an unused candidate. When a surname matches >1 distinct
    candidate, first try the city carried in `display`; if still ambiguous, defer (return None) unless
    `allow_ambiguous`, in which case pick deterministically (first by USTA ID)."""
    s, trunc = _core(surname)
    avail = [p for p in cands if p.usta_id not in used]
    hit = [p for p in avail if _norm(p.last) == s]
    if len(hit) == 1:
        return hit[0], ""
    if len(hit) > 1:                               # same-surname collision
        narrowed = _city_pick(hit, display)
        if len(narrowed) == 1:
            return narrowed[0], f"city-matched '{surname}' -> {narrowed[0].name}"
        if not allow_ambiguous:
            return None, ""                        # defer to the forcing pass
        hit.sort(key=lambda p: p.usta_id)
        return hit[0], f"same-surname '{surname}' -> {hit[0].name}"
    # surname carries a city tail ("Shooshtarian San Anselmo") or was truncated ("Clevle...")
    lead = [p for p in avail if s.startswith(_norm(p.last) + " ") or _norm(p.last) == s.split(" ")[0]
            or (trunc and _norm(p.last).startswith(s))]
    if len(lead) > 1:
        lead = _city_pick(lead, display) or lead
    if lead:
        lead.sort(key=lambda p: -len(p.last))
        return lead[0], f"{'truncated' if trunc else 'partial'}-matched '{surname}' -> {lead[0].name}"
    return None, f"no last-name match '{surname}'"


def _resolve_entrant(division, display, seed, partners, is_doubles, cands, used, kind,
                     pos=None, is_bye=False, allow_ambiguous=True):
    """Resolve one entrant; `used` is the division-level set of already-claimed player ids. When
    `allow_ambiguous` is False, an entrant whose surname can't be pinned yet is left unresolved
    (and claims nothing) so a later forcing pass can settle it once the certain ones are consumed."""
    r = Resolved(division=division, kind=kind, pos=pos, display=display, seed=seed, is_bye=is_bye)
    if is_bye:
        r.ok = True
        return r
    if is_doubles:
        ids, notes = [], []
        for sur in partners:
            p, note = _match_last(sur, cands, used, display=display, allow_ambiguous=allow_ambiguous)
            if p:
                ids.append(p.usta_id)
                used.add(p.usta_id)      # provisional; caller rolls back if the entrant is deferred
            if note:
                notes.append(note)
        r.player_ids = ids
        r.ok = len(ids) == 2
        r.note = "; ".join(n for n in notes if n)
        if not r.ok and not r.note:
            r.note = f"doubles '{display}' resolved {len(ids)}/2"
    else:
        p, note = _match_singles(display, cands, used)
        if p:
            r.player_ids = [p.usta_id]
            used.add(p.usta_id)
            r.ok = True
        r.note = note
    return r


def _entrant_descs(d):
    """Flatten a DivisionDraw into entrant descriptors in draw order (ref, kind, partners, ...)."""
    descs = []
    if d.fmt == "single_elim":
        for s in d.slots:
            descs.append({"ref": f"slot-{s.pos}", "kind": "slot", "pos": s.pos, "group": "",
                          "display": s.display, "seed": s.seed, "partners": s.partners,
                          "is_doubles": s.is_doubles, "is_bye": s.is_bye})
    else:
        for g in d.groups:
            for i, m in enumerate(g.members):
                descs.append({"ref": f"{g.name}#{i}", "kind": "rr", "pos": None, "group": g.name,
                              "display": m.display, "seed": m.seed, "partners": m.partners,
                              "is_doubles": m.is_doubles, "is_bye": False})
    return descs


def _resolve_division(event, descs, cands):
    """Two-pass resolution: pass 1 settles only the certain entrants (deferring any surname that
    can't be pinned yet, rolling back its provisional claims); pass 2 forces the rest once the
    certain players are consumed. Draw order is preserved in the returned list."""
    used, results = set(), {}
    for allow in (False, True):
        for e in descs:
            if e["ref"] in results and results[e["ref"]].ok:
                continue
            snapshot = set(used)
            r = _resolve_entrant(event, e["display"], e["seed"], e["partners"], e["is_doubles"],
                                 cands, used, e["kind"], pos=e["pos"], is_bye=e["is_bye"],
                                 allow_ambiguous=allow)
            r.ref = e["ref"]; r.group = e["group"]; r.is_doubles = e["is_doubles"]
            if r.ok or e["is_bye"] or allow:
                results[e["ref"]] = r                       # commit
            else:
                used.clear(); used.update(snapshot)         # defer: undo provisional claims
    return [results[e["ref"]] for e in descs]


def resolve_draws(draws, players):
    """Resolve every drawn entrant to canonical player(s). Returns (resolved_by_division, stats)."""
    didx = _division_index(players)
    out = {}
    total = ok = 0
    for d in draws:
        cands = didx.get(d.event, list(players.values()))   # fall back to global pool
        res = _resolve_division(d.event, _entrant_descs(d), cands)
        for r in res:
            if not r.is_bye:
                total += 1
                ok += r.ok
        out[d.event] = res
    stats = {"entrants": total, "resolved": ok, "unresolved": total - ok,
             "rate": round(ok / total, 4) if total else 0.0}
    return out, stats


def roster_meta(players):
    """{name: {"city","section","zip"}} for the locality re-base (constraints.local_players_from_
    locality). Real export has no zip → zip is "" and locality runs off city/section."""
    return {p.name: {"city": p.city, "section": p.section, "zip": ""} for p in players.values()}


def _is_doubles(division):
    return "doubles" in division.lower() or "mixed" in division.lower()


# ---- R1d: TD ingest-review projection + overrides ---------------------------

INGEST_REVIEW_SCHEMA = "td-ingest-review/v1"
INGEST_OVERRIDES_SCHEMA = "td-ingest-review-overrides/v1"


def ingest_review_plan(level="2", td_path=None, st_path=None, draws_path=None):
    """Self-contained `td-ingest-review/v1` for the TD review console: every drawn entrant with its
    resolved player(s), seed, bye, and confidence (exact | flagged + note), plus each division's
    candidate roster for the override pickers. Read-only; couriered to the console (B-1)."""
    players = load_players(td_path, st_path, level)
    draws = draws_pdf.parse_draws(draws_path, level=level)
    by_div, stats = resolve_draws(draws, players)
    dmap = {d.event: d for d in draws}
    didx = _division_index(players)

    def pref(pid):
        p = players.get(pid)
        return {"usta_id": pid, "name": p.name if p else pid}

    divisions = []
    for div, res in by_div.items():
        d = dmap[div]
        entrants = []
        for r in res:
            if r.is_bye:
                entrants.append({"ref": r.ref, "pos": r.pos, "group": r.group, "is_bye": True,
                                 "is_doubles": r.is_doubles, "seed": None, "display": "BYE",
                                 "players": [], "confidence": "exact", "note": ""})
                continue
            entrants.append({
                "ref": r.ref, "pos": r.pos, "group": r.group, "is_bye": False,
                "is_doubles": r.is_doubles, "seed": r.seed, "display": r.display,
                "players": [pref(i) for i in r.player_ids],
                "confidence": "flagged" if (r.note or not r.ok) else "exact",
                "note": r.note if r.ok else (r.note or "unresolved"),
            })
        cands = sorted(({"usta_id": p.usta_id, "name": p.name, "city": p.city,
                         "section": p.section} for p in didx.get(div, [])),
                       key=lambda c: c["name"])
        divisions.append({"event": div, "fmt": d.fmt, "is_doubles": _is_doubles(div),
                          "entrants": entrants, "candidates": cands})
    return {"schema": INGEST_REVIEW_SCHEMA, "level": level,
            "tournament": "USTA Wilson World Tennis Classic",
            "resolution": stats, "divisions": divisions}


def apply_review_overrides(by_div, overrides_doc):
    """Apply a `td-ingest-review-overrides/v1` doc onto a resolution ({event: [Resolved]}), in place.
    Each override targets an entrant by (event, ref) and may set players (usta_ids), seed, is_bye.
    A TD override is authoritative → marks the entrant resolved. Returns the count applied."""
    if not overrides_doc:
        return 0
    if overrides_doc.get("schema") not in (INGEST_OVERRIDES_SCHEMA, None):
        raise ValueError(f"unexpected overrides schema {overrides_doc.get('schema')!r}")
    index = {(ev, r.ref): r for ev, res in by_div.items() for r in res}
    applied = 0
    for ov in overrides_doc.get("overrides", []):
        key = (ov.get("event"), ov.get("ref"))
        r = index.get(key)
        if r is None:
            raise ValueError(f"override targets unknown entrant {key}")
        if "is_bye" in ov:
            r.is_bye = bool(ov["is_bye"])
        if "players" in ov:
            r.player_ids = [str(x) for x in ov["players"]]
            r.ok = bool(r.player_ids) and not r.is_bye
        if "seed" in ov:
            r.seed = ov["seed"]
        r.note = (r.note + "; " if r.note else "") + "TD override"
        applied += 1
    return applied


def division_parent(name):
    """The parent division of an EventSpec name. Round-robin groups are minted as
    `f"{div} — {g.name}"` in `load_from_finalized_draws` (one EventSpec per printed group), while
    the TD list only ever names the parent. Every join between the two sides must go through here
    — comparing the raw names classified **every** round-robin entrant as undrawn, which is what
    made 55 of 108 non-drawn cards false and printed "Everyone entered in this division is in the
    draw" on 9 group panels. (Absorbs OI-31, whose recorded root cause named `_expand_rr_groups`;
    the suffix is minted here in the ingest, not there.)"""
    return str(name).split(" — ")[0]


def _norm_div(s):
    return "".join(c for c in str(s).lower() if c.isalnum())


def reconcile_entries(players, events, schedule=None, draws=None, ingest_warnings=None):
    """**The closed accounting.** Every person-division entry on the TD lists lands in exactly one
    of four buckets, and the four sum to the entry count — nothing can go missing without the
    total failing. This is what lets the tool answer "did everyone who should be playing get
    placed?", which 0-unplaced never did: 0 unplaced only means every match the tool BUILT, it
    placed, and it never compared the entry lists to the schedule at all.

    Buckets, in priority order:
      - **scheduled** — the person is in the division's Raw Draw. Per the Director's ruling
        (2026-07-27) *the draw is truth*: anyone in the Raw Draw is in, whatever the entry list
        says, including the 45 withdrawn-but-drawn players the desk kept. Status is never consulted
        here.
      - **not_buildable** — the person is in the Raw Draw, but the ingest could not build the
        structure that holds them (an unresolvable teammate collapsed the slot or the group), and
        an ingest warning says so. PREP-1 (2026-08-01): entrants of an ingest-dropped structure
        used to fall into `no_draw` with reason text asserting the division *"printed no draw"* —
        false on its face, the TD is holding that draw. The row's reason carries the ingest's own
        warning verbatim.
      - **not_drawn** — the division HAS a Raw Draw and this person is not in it.
      - **no_draw** — the division never printed a draw at all.

    The last three together are the **exceptions list**: entered, not playing, with the desk's own
    words (or the ingest's own warning) attached. That is information, not a failure. The failure
    cases are its mirror, returned as `drawn_not_scheduled` for the caller to assert on
    (`wwtc_pipeline` raises; it is 0 on the reference field): a person in the Raw Draw with no
    match on the schedule, and — PREP-1's sharp rule, **every lost entrant is either warned about
    or fatal, never silent** — a drawn entrant the ingest lost with NO warning covering their
    division. A warned loss is `not_buildable`; an unwarned loss stays fatal.

    `draws` (PREP-1, A1) is the PARSED-draw truth: `{division: [drawn player names]}` off
    `draws_pdf.parse_draws` + `resolve_draws` (`load_from_finalized_draws` returns it as
    `meta["drawn_by_division"]`; `draw_truth` merges levels). It is what `has_draw` and "drawn"
    are built from — the post-drop `events` list confirmed its own drops, which is how a deleted
    division's entrants were mislabelled. `ingest_warnings` is `meta["warnings"]`, joined per
    division for the warned-or-fatal rule. When `draws` is None (synthetic/unit callers with no
    PDFs behind them), both fall back to the built events — that path cannot see an ingest drop
    and never emits `not_buildable`.

    `events` is the built EventSpec list (round-robin groups included — they are joined through
    `division_parent`). `schedule` is the finished `result["schedule"]`; omit it to reconcile
    against the draws alone, in which case the schedule half of `drawn_not_scheduled` is not
    computed (the unwarned-ingest-loss half needs no schedule and always is).

    Identity is the engine's human key (`First Last`), the same key `Team.members` and the rest
    rules use, so a person the engine treats as one human is one person here too.
    Deterministic: every list is sorted.
    """
    built = {}
    for ev in events:
        parent = division_parent(ev.name)
        for t in ev.teams:
            for who in (t.members or []):
                if who:
                    built.setdefault(who, set()).add(_norm_div(parent))

    if draws is not None:
        has_draw = {_norm_div(div): div for div in draws}
        drawn = {}
        for div, names in draws.items():
            for who in names:
                drawn.setdefault(who, set()).add(_norm_div(div))
    else:
        has_draw = {}
        for ev in events:
            parent = division_parent(ev.name)
            has_draw[_norm_div(parent)] = parent
        drawn = built

    # The warned-or-fatal join: an ingest warning covers a division when its leading token names
    # that division (every `load_from_finalized_draws` warning starts `<division or group>: …`).
    warned = {}
    for w in (ingest_warnings or ()):
        warned.setdefault(_norm_div(division_parent(str(w).split(":", 1)[0])), []).append(str(w))

    played = {}
    if schedule is not None:
        for row in schedule:
            key = _norm_div(division_parent(row.get("event", "")))
            for who in row.get("players", []) or []:
                played.setdefault(who, set()).add(key)

    # DESK-1: who each person is standing on a court WITH, from the built teams. Feeds D6's
    # Unpaired-but-drawn check only — it is read, never placed.
    partners = {}
    for ev in events:
        pkey = _norm_div(division_parent(ev.name))
        for t in ev.teams:
            mem = [m for m in (t.members or []) if m]
            for who in mem:
                mate = sorted(m for m in mem if m != who)
                if mate:
                    partners[(who, pkey)] = mate[0]

    scheduled, not_drawn, no_draw, not_buildable, drawn_not_scheduled = [], [], [], [], []
    unpaired_but_drawn = []
    for p in sorted(players.values(), key=lambda x: (x.name, x.usta_id)):
        # DESK-1: the draw-informed reading, computed ONCE per person against the same draw
        # membership the buckets below are built from, so the label and the reason can never
        # disagree. The Player is not mutated — `status_by_div` is this call's own answer.
        mine = drawn.get(p.name, set())
        status_by_div = resolved_entry_status(p, mine)
        mates = unpaired_partners(p, mine, partners)
        for division in p.entry_status:
            key = _norm_div(division)
            st, mate = status_by_div.get(division), mates.get(division)
            if key in drawn.get(p.name, ()):
                label = has_draw.get(key, division)
                if key in built.get(p.name, ()):
                    row = _entry_row(p, division, "scheduled", label, status=st, partner=mate,
                                     row_status=status_by_div)
                    scheduled.append(row)
                    if mate:
                        unpaired_but_drawn.append(row)
                    if schedule is not None and key not in played.get(p.name, ()):
                        drawn_not_scheduled.append(row)
                elif key in warned:
                    not_buildable.append(_entry_row(p, division, "not_buildable", label,
                                                    detail="; ".join(warned[key]), status=st,
                                                    row_status=status_by_div))
                else:
                    # Drawn per the Raw Draw, absent from the built events, and no ingest
                    # warning covers the division: an unwarned loss. Never silent — fatal.
                    row = _entry_row(p, division, "scheduled", label, status=st, partner=mate,
                                     row_status=status_by_div)
                    scheduled.append(row)
                    if mate:
                        unpaired_but_drawn.append(row)
                    drawn_not_scheduled.append(row)
            elif key in has_draw:
                not_drawn.append(_entry_row(p, division, "not_drawn", has_draw[key], status=st,
                                          row_status=status_by_div))
            else:
                no_draw.append(_entry_row(p, division, "no_draw", division, status=st,
                                         row_status=status_by_div))

    exceptions = sorted(not_drawn + no_draw + not_buildable,
                        key=lambda r: (r["division"], r["name"], r["usta_id"]))
    entries = len(scheduled) + len(exceptions)
    if entries != sum(len(p.entry_status) for p in players.values()):
        raise AssertionError("reconciliation is not closed: buckets do not sum to the entry count")

    by_event = {}
    for r in exceptions:
        by_event.setdefault(r["division"], []).append(r)
    return {
        "entries": entries,
        "counts": {"entries": entries, "scheduled": len(scheduled),
                   "not_drawn": len(not_drawn), "no_draw": len(no_draw),
                   "not_buildable": len(not_buildable),
                   "exceptions": len(exceptions),
                   "unpaired_but_drawn": len(unpaired_but_drawn)},
        "people": {"entered": len({p.name for p in players.values() if p.entry_status}),
                   "exceptions": len({r["name"] for r in exceptions})},
        "scheduled": scheduled,
        "exceptions": exceptions,
        "by_event": {k: by_event[k] for k in sorted(by_event)},
        "unclear": sum(1 for r in exceptions if r["unclear"]),
        "unclear_people": len({r["name"] for r in exceptions if r["unclear"]}),
        "drawn_not_scheduled": drawn_not_scheduled,
        # DESK-1 (run report D6): entries the desk marked `Unpaired` that are standing in that
        # division's Raw Draw with a partner. Every one of them is IN a draw, so they are
        # `scheduled` rows and never reach the exceptions list — this is the list that makes
        # them countable, and each row's own `reason` names both sources.
        "unpaired_but_drawn": unpaired_but_drawn,
    }


def _entry_row(p, division, bucket, label, detail=None, status=_UNSET, partner=None,
               row_status=None):
    """One reconciliation row. `label` is the DRAW's spelling of the division when it ran (so the
    report and the draw sheets agree), else the TD list's own string. `detail` is the ingest's
    own warning text for a `not_buildable` row — the reason column carries it verbatim.

    `status` (DESK-1) is the draw-informed reading for this entry; omit it and the row falls back
    to `p.entry_status`, which is what a caller with no draws behind it still gets. `partner` is
    the person this entry is drawn WITH when the desk marked it `Unpaired` — D6's check.

    **EVERY SENTENCE HERE IS FACT -> EXPLANATION -> ACTION** (8/7 note 9). The reason column is
    read by a director deciding what to do about a name, and the pre-DESK-1 strings stopped at
    the fact: `"Alternate"` on its own, or `"Withdrawn — division printed no draw"`, told him a
    word and left him to work out whether it was his problem. Where there is nothing to do, the
    sentence says so; where there is, it says what.

    **Vocabulary: Raw Draw** — the ingested Tournament Desk export, the truth (`LANG-1_glossary`
    §2a, ruled 2026-08-09). The phrase these sentences used to carry is retired: it named both
    the desk's export and the tool's own rendered draw surfaces and distinguished neither, and
    this column is about exactly one of them.
    """
    row = p.raw_td_by_division.get(division) or p.raw_td
    desk = str((row or {}).get("Draw status") or p.td_status or "").strip()
    if status is _UNSET:
        status = p.entry_status.get(division)
    unclear = status is None
    settled = status == NO_DESK_WORD                 # resolved, but the desk gave no reason
    if bucket == "not_buildable":
        # PREP-1: the person IS in the Raw Draw (draws are truth — status is irrelevant here),
        # and the tool, not the desk, is why they have no schedule. Say so, in the ingest's own
        # words, and point at the surface that can settle it.
        reason = (f"In the Raw Draw, but the tool could not build the structure that holds this "
                  f"entry: {detail}. The desk's draw stands — check that page of the Raw Draw.")
    elif unclear:
        # Operator ruling (b), 2026-07-31: print the raw row, invent nothing. The row's own
        # divisions are named so the TD can see exactly which cell is ambiguous — the ambiguity
        # belongs to one TD ROW, not to the player's other level, which has its own cell.
        # LANG-1 item 46: fixed HERE, at the source. This one string feeds three surfaces — the
        # Edit console's panel, the withdrawn list and the exceptions CSV — so patching it at a
        # render site would leave the other two reading the old words.
        # DESK-1: the row's STILL-UNRESOLVED divisions, not all of them. The draw settles part of
        # a conflicted row — Harmon's Mixed 40 & over doubles is one of his row's four entries and
        # he is standing in that draw — and naming it here would ask the desk to split something
        # already split. `row_status` is this call's reading; without it (a caller with no draws)
        # the whole row is named, exactly as before.
        siblings = sorted(d for d, r in p.raw_td_by_division.items()
                          if r == row and (row_status or {}).get(d, None) is None)
        reason = (f'Entry list conflicts — {desk.replace(", ", " / ")}: ' + ", ".join(siblings)
                  + ". The entry list gives more than one non-playing status for this row and "
                    "nothing says which division each belongs to. Ask the desk to split this "
                    "row's status per division.") if desk else \
            ("The entry list gives no draw status for this row, so there is no word to attach "
             "to any of these entries. Ask the desk to add one.")
    elif bucket == "scheduled":
        reason = "In the Raw Draw."
    elif settled:
        where = ("this division never printed a draw" if bucket == "no_draw"
                 else "not in the Raw Draw for this division")
        # True of both ways an entry gets here: the desk's word for it was `Selected`, or the
        # row's one word was and it covers several entries. Either way the entry list says the
        # player was picked and never says why they are out, and the Raw Draw disagrees.
        reason = (f"Not playing — {where}. The entry list records no reason: the status it "
                  f"carries says the player was selected, not why they are out. Ask the desk "
                  f"to confirm it.")
    elif bucket == "no_draw":
        reason = (f"{status} on the entry list — this division never printed a draw, so nobody "
                  f"entered in it is playing. Nothing to do here.")
    else:
        reason = (f"{status} on the entry list — not in the Raw Draw for this division. "
                  f"Nothing to do unless the desk expected them to play.")
    if partner:
        # D6, 2026-08-09. Both sources named, and the question put to the only party who can
        # answer it. This rides on a SCHEDULED row: the person is in the draw, which is the
        # whole point — the desk's word and the draw cannot both be current.
        reason += (f" Check — the entry list marks this entry Unpaired, but the Raw Draw has "
                   f"them playing with {partner}. Ask the desk which is stale.")
    return {"name": p.name, "usta_id": p.usta_id, "division": label, "bucket": bucket,
            # LANG-1 item 46's family: this default is a DISPLAY string too — it reaches the Edit
            # console's panel behind a click, which is where the ship sweep caught it (529 hits
            # a source-only reading would have shipped). DESK-1 (8/7 note 9): "Not stated" said
            # nothing about WHO had not stated it; the entry list is the thing that is silent.
            "status": "No status on the entry list" if (unclear or settled) else status,
            "reason": reason, "desk_status": desk, "unclear": unclear,
            # DESK-1 (§2, Operator 8/8): `status == "Selected"` is RETIRED as a review trigger.
            # Under the draw-informed reading every drawn entry reads `Selected`, so keeping it
            # would flag all 1,066 scheduled rows — it fired on ~1 row before and would have
            # re-flagged every broadcast-Selected row. A settled row is NOT flagged either: the
            # fact is known, only the desk's reason is missing. What is left is a genuine
            # ambiguity, a structure the tool could not build, and D6's two-source conflict.
            "review": unclear or bucket == "not_buildable" or bool(partner)}


def non_drawn_entrants(players, events, draws=None, ingest_warnings=None):
    """F6 — the entered-but-not-playing pool, as the edit console's panel consumes it.

    **ROSTER-1 (2026-07-31): this is now a projection of `reconcile_entries`, not a second
    computation.** It used to derive its own answer, and the divergence between the two is exactly
    what produced today's false cards: it compared the TD list's parent division names against the
    engine's `"<parent> — Group N"` group names, so every round-robin entrant looked undrawn —
    **55 of 108 cards false**, 8 divisions wrongly listed as having no draw, and 9 group panels
    affirmatively printing "Everyone entered in this division is in the draw." The panel and the
    exceptions list now read the same buckets and cannot disagree again.

    `draws` / `ingest_warnings` (PREP-1) pass straight through to `reconcile_entries` — the
    parsed-draw truth and the warned-or-fatal join. A `not_buildable` card always lands in
    `by_event` (never `withdrawn`, whatever the desk's status string says — the person is in the
    Raw Draw) and carries its bucket so the panel can show the could-not-build reason.

    Surface + flag ONLY — nothing here is ever placed, and status is never a placement input.
    ONE bounded exception (DRAW-1, 2026-08-06, ruling 84): a `substitute` edit may bring a
    person from this pool INTO AN EXISTING MATCH as a doubles partner. That is the whole
    exception — these people are still never placed into a slot of their own, the panel is
    still surface + flag, and status is still never a placement input.

    Returns {"by_event": {division: [card]}, "withdrawn": [card], "ambiguous": n} where card =
    {name, usta_id, event, status, review, reason, desk_status, unclear, bucket}. `event` is the
    DRAW's division name when the division ran, else the TD list's string. Deterministic.
    """
    rec = reconcile_entries(players, events, draws=draws, ingest_warnings=ingest_warnings)
    by_event, withdrawn = {}, []
    for r in rec["exceptions"]:
        card = {"name": r["name"], "usta_id": r["usta_id"], "event": r["division"],
                "status": r["status"], "review": r["review"], "reason": r["reason"],
                "desk_status": r["desk_status"], "unclear": r["unclear"],
                "bucket": r["bucket"]}
        if r["bucket"] != "not_buildable" and not r["unclear"] \
                and "withdrawn" in str(r["status"]).lower():
            withdrawn.append(card)
        else:
            by_event.setdefault(card["event"], []).append(card)
    for v in by_event.values():
        v.sort(key=lambda c: (c["name"], c["usta_id"]))
    withdrawn.sort(key=lambda c: (c["event"], c["name"]))
    return {"by_event": dict(sorted(by_event.items())), "withdrawn": withdrawn,
            "ambiguous": rec["unclear"]}


def load_from_finalized_draws(level="2", td_path=None, st_path=None, draws_path=None,
                              overrides=None):
    """Ingest the real finalized draw → (events, seeds_by_event, meta), a drop-in for the engine
    the way serve_tennis_intake.load_export is: `schedule_multi(config_from_slate(slate, events))`.

    Elim division → one EventSpec(fmt="single_elim") whose `teams` is the ordered bracket with
    BYE sentinels (auto-routes to scheduler_multi._build_elim_positional — verbatim pairing, no
    re-permutation). RR division → one EventSpec(fmt="round_robin") per printed group. Doubles are
    2-member Teams. Seeds ride `seeds_by_event` (render-only). The engine is not modified.
    """
    from scheduler_multi import EventSpec, Team, BYE
    from serve_tennis_intake import _recovery_for, _cap_for

    players = load_players(td_path, st_path, level)
    draws = draws_pdf.parse_draws(draws_path, level=level)
    by_div, stats = resolve_draws(draws, players)
    apply_review_overrides(by_div, overrides) if overrides else None
    dmap = {d.event: d for d in draws}

    # PREP-1 (A1): the printed-draw truth `reconcile_entries` joins against — every PARSED
    # division, with the names its resolution could attach (a partly-resolved doubles team's
    # resolved partner included: their name is on the printed page). Computed here, one layer
    # below the structure-building that can drop a slot or a whole division, so the post-drop
    # `events` list can never again vouch for what the desk printed.
    drawn_by_division = {
        div: sorted({players[pid].name for r in res for pid in r.player_ids if pid in players})
        for div, res in by_div.items()}

    def team_for(r):
        names = [players[i].name for i in r.player_ids]
        return Team(tid="+".join(sorted(r.player_ids)), members=names)

    events, seeds_by_event, warnings = [], {}, []
    for div, res in by_div.items():
        d = dmap[div]
        doubles = _is_doubles(div)
        common = dict(match_minutes=90, recovery_minutes=_recovery_for(div),
                      precedence=1 if doubles else 0, max_matches_per_day=_cap_for(div))
        if d.fmt == "single_elim":
            teams, seedmap = [], {}
            for r in res:
                if r.is_bye:
                    teams.append(Team(tid=BYE, members=[]))
                elif r.ok:
                    t = team_for(r)
                    teams.append(t)
                    if r.seed:
                        seedmap[t.label()] = r.seed
                else:
                    teams.append(Team(tid=BYE, members=[]))   # unresolved → treat as bye (logged)
                    warnings.append(f"{div}: unresolved slot {r.pos} {r.display!r} -> BYE")
            events.append(EventSpec(name=div, fmt="single_elim", teams=teams, **common))
            if seedmap:
                seeds_by_event[div] = seedmap
        else:  # round_robin — one EventSpec per group
            for g in d.groups:
                gname = f"{div} — {g.name}"
                teams, seedmap = [], {}
                for m in g.members:
                    r = next((x for x in res if x.display == m.display and x.kind == "rr"), None)
                    if r and r.ok:
                        t = team_for(r)
                        teams.append(t)
                        if r.seed:
                            seedmap[t.label()] = r.seed
                    else:
                        warnings.append(f"{gname}: unresolved member {m.display!r} -> dropped")
                if len(teams) >= 3:
                    events.append(EventSpec(name=gname, fmt="round_robin", teams=teams, **common))
                    if seedmap:
                        seeds_by_event[gname] = seedmap
                else:
                    warnings.append(f"{gname}: only {len(teams)} teams (<3) — skipped")
    meta = {"players": len(players), "resolution": stats, "warnings": warnings,
            "n_events": len(events), "drawn_by_division": drawn_by_division}
    return events, seeds_by_event, meta


def draw_truth(metas):
    """PREP-1: merge per-level ingest metas into the `(draws, ingest_warnings)` pair
    `reconcile_entries` takes — the parsed-draw truth unioned across levels (division names never
    collide across levels; `build_combined` asserts it) plus every ingest warning, in order."""
    drawn, warnings = {}, []
    for meta in metas:
        for div, names in (meta.get("drawn_by_division") or {}).items():
            drawn.setdefault(div, set()).update(names)
        warnings.extend(meta.get("warnings") or [])
    return {div: sorted(v) for div, v in drawn.items()}, warnings


def _selftest():
    td, st = resolve_player_lists("2")
    print(f"TD list: {os.path.basename(td)}\nST list: {os.path.basename(st)}")
    players = load_players(td, st)
    with_sec = sum(1 for p in players.values() if p.section)
    print(f"players joined: {len(players)}  with Section: {with_sec}")
    draws = draws_pdf.parse_draws()
    by_div, stats = resolve_draws(draws, players)
    print(f"draw entrants: {stats['entrants']}  resolved: {stats['resolved']}  "
          f"unresolved: {stats['unresolved']}  rate: {stats['rate']*100:.1f}%")
    # show a few unresolved for review
    shown = 0
    for div, res in by_div.items():
        for r in res:
            if not r.ok and not r.is_bye:
                print(f"  UNRESOLVED [{div}] {r.display!r} -> {r.note}")
                shown += 1
        if shown >= 25:
            break
    print("SELF-TEST OK" if stats["rate"] > 0 else "SELF-TEST: no resolution")


if __name__ == "__main__":
    _selftest()
